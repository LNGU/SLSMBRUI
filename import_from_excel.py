#!/usr/bin/env python3
"""
Import SLS MBR data from an Excel workbook into data.js.

Reads sheets from an Excel file, maps columns to the data model,
and updates the defaultRawData object in data.js.

Sheet detection (case-insensitive partial match):
  - "publisher"           → publishers array
  - "spend"               → spendData array
  - "risk"                → riskData array
  - "title" or "managed"  → managedTitles array
  - "kpi" or "external"   → externalKpis array

Column headers are auto-mapped by fuzzy matching (case-insensitive).
Unrecognized sheets/columns are skipped with a warning.

Usage:
  python import_from_excel.py --file data.xlsx
  python import_from_excel.py --file data.xlsx --dry-run
  python import_from_excel.py --file data.xlsx --sheet-map "Sheet1=publishers" "Sheet2=spend"
  python import_from_excel.py --template output.xlsx

Prerequisites:
  pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

if not openpyxl and not xlrd:
    print("ERROR: Install openpyxl (for .xlsx) and/or xlrd (for .xls): pip install openpyxl xlrd", file=sys.stderr)
    sys.exit(1)


# ─── Column Mappings ─────────────────────────────────────────────────────────
# Each entry: list of possible header names → target field name
# First match wins (case-insensitive substring match)

PUBLISHER_COLUMNS = [
    (["id", "publisher id"], "id"),
    (["name", "publisher name", "publisher"], "name"),
    (["title", "product", "product title", "software"], "title"),
    (["type", "license type", "lic type"], "type"),
    (["contact", "owner", "manager"], "contact"),
    (["renewal date", "renewal", "renew date", "expiry", "expiration"], "renewalDate"),
    (["status"], "status"),
    (["savings amount", "savings", "saving"], "savingsAmount"),
    (["savings type", "saving type"], "savingsType"),
]

SPEND_COLUMNS = [
    (["publisher", "name", "publisher name"], "publisher"),
    (["company spend", "company", "total spend", "org spend"], "companySpend"),
    (["msd spend", "msd", "ms dev"], "msdSpend"),
    (["tiam spend", "ti&m spend", "ti&m", "tiam", "ti and m"], "tiamSpend"),
    (["fiscal year", "fy", "year"], "fiscalYear"),
    (["notes", "note", "comments"], "notes"),
]

RISK_COLUMNS = [
    (["publisher", "name", "publisher name"], "publisher"),
    (["sspa"], "sspa"),
    (["po", "purchase order"], "po"),
    (["finance", "financial"], "finance"),
    (["legal"], "legal"),
    (["inventory", "inv"], "inventory"),
    (["details", "detail", "notes", "comments"], "details"),
]

TITLE_COLUMNS = [
    (["title", "managed title", "product", "software title"], "title"),
    (["publisher", "publisher name", "vendor"], "publisher"),
    (["category", "cat"], "category"),
    (["license count", "count", "qty", "quantity", "licenses"], "licenseCount"),
    (["notes", "note", "status", "comments"], "notes"),
]

KPI_COLUMNS = [
    (["name", "kpi", "kpi name", "metric"], "name"),
    (["value", "val", "amount", "count"], "value"),
    (["unit", "uom"], "unit"),
    (["source", "system", "data source"], "source"),
    (["last updated", "updated", "date", "as of"], "lastUpdated"),
    (["notes", "note", "comments"], "notes"),
]

SHEET_TYPE_MAPPINGS = {
    "publishers": PUBLISHER_COLUMNS,
    "spend": SPEND_COLUMNS,
    "risks": RISK_COLUMNS,
    "titles": TITLE_COLUMNS,
    "kpis": KPI_COLUMNS,
}


# ─── Sheet Detection ─────────────────────────────────────────────────────────

def detect_sheet_type(sheet_name: str) -> Optional[str]:
    """Auto-detect which data type a sheet contains based on its name."""
    lower = sheet_name.lower().strip()
    if "publisher" in lower and "managed" not in lower:
        return "publishers"
    if "spend" in lower:
        return "spend"
    if "risk" in lower:
        return "risks"
    if "title" in lower or "managed" in lower:
        return "titles"
    if "kpi" in lower or "external" in lower:
        return "kpis"
    return None


# ─── Column Mapping ──────────────────────────────────────────────────────────

def map_columns(headers: List[str], column_defs: list) -> Dict[int, str]:
    """Map Excel column indices to data field names via fuzzy matching."""
    mapping: Dict[int, str] = {}
    used_fields = set()

    for col_idx, header in enumerate(headers):
        if not header:
            continue
        h = header.strip().lower()
        for aliases, field_name in column_defs:
            if field_name in used_fields:
                continue
            # Exact match first
            if h in [a.lower() for a in aliases]:
                mapping[col_idx] = field_name
                used_fields.add(field_name)
                break
            # Substring match
            for alias in aliases:
                if alias.lower() in h or h in alias.lower():
                    mapping[col_idx] = field_name
                    used_fields.add(field_name)
                    break
            if col_idx in mapping:
                break

    return mapping


# ─── Data Reading ─────────────────────────────────────────────────────────────

def read_sheet_data(ws, column_defs: list) -> Tuple[List[Dict], Dict[int, str], List[str]]:
    """Read an openpyxl worksheet into a list of dicts using column mapping."""
    rows = list(ws.iter_rows(values_only=True))
    return read_sheet_data_from_rows(rows, column_defs)


def read_sheet_data_from_rows(rows: List[tuple], column_defs: list) -> Tuple[List[Dict], Dict[int, str], List[str]]:
    """Read raw row tuples into a list of dicts using column mapping."""
    if not rows:
        return [], {}, []

    # Find header row (first non-empty row)
    header_row_idx = 0
    for i, row in enumerate(rows):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            header_row_idx = i
            break

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[header_row_idx]]
    col_mapping = map_columns(headers, column_defs)

    unmapped = [h for i, h in enumerate(headers) if h and i not in col_mapping]

    records = []
    for row in rows[header_row_idx + 1:]:
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue  # skip empty rows

        record = {}
        for col_idx, field_name in col_mapping.items():
            val = row[col_idx] if col_idx < len(row) else None
            record[field_name] = convert_value(val, field_name)

        # Skip rows where the primary key field is empty
        primary = record.get("name") or record.get("publisher") or record.get("title")
        if primary:
            records.append(record)

    return records, col_mapping, unmapped


# ─── Format Readers ───────────────────────────────────────────────────────────

def _is_old_xls_format(file_path: Path) -> bool:
    """Check if the file is an old .xls format by reading its magic bytes."""
    with open(file_path, "rb") as f:
        header = f.read(8)
    # OLE2 Compound Binary File magic: D0 CF 11 E0 A1 B1 1A E1
    return header[:4] == b"\xd0\xcf\x11\xe0"


def _read_xlsx(file_path: Path) -> Tuple[List[str], Dict[str, List[tuple]]]:
    """Read an .xlsx file via openpyxl, return sheet names and rows per sheet."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    sheet_rows: Dict[str, List[tuple]] = {}
    for name in sheet_names:
        ws = wb[name]
        sheet_rows[name] = list(ws.iter_rows(values_only=True))
    wb.close()
    return sheet_names, sheet_rows


def _read_xls(file_path: Path) -> Tuple[List[str], Dict[str, List[tuple]]]:
    """Read an old .xls file via xlrd, return sheet names and rows per sheet."""
    wb = xlrd.open_workbook(str(file_path), formatting_info=False)
    sheet_names = wb.sheet_names()
    sheet_rows: Dict[str, List[tuple]] = {}
    for name in sheet_names:
        ws = wb.sheet_by_name(name)
        rows = []
        for row_idx in range(ws.nrows):
            row_vals = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                val = cell.value
                # Convert xlrd date numbers to Python dates
                if cell.ctype == xlrd.XL_CELL_DATE and val:
                    try:
                        dt_tuple = xlrd.xldate_as_tuple(val, wb.datemode)
                        val = datetime(*dt_tuple)
                    except Exception:
                        pass
                row_vals.append(val)
            rows.append(tuple(row_vals))
        sheet_rows[name] = rows
    return sheet_names, sheet_rows


def convert_value(val: Any, field_name: str) -> Any:
    """Convert an Excel cell value to the appropriate Python type."""
    if val is None:
        # Return sensible defaults based on field
        if field_name in ("savingsAmount", "companySpend", "msdSpend", "tiamSpend", "licenseCount", "value"):
            return 0
        if field_name == "id":
            return 0
        return ""

    # Dates
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")

    # Numbers
    if field_name in ("savingsAmount", "companySpend", "msdSpend", "tiamSpend", "value"):
        try:
            return float(val) if val else 0
        except (ValueError, TypeError):
            return 0
    if field_name in ("id", "licenseCount"):
        try:
            return int(float(val)) if val else 0
        except (ValueError, TypeError):
            return 0

    # Strings
    return str(val).strip() if val else ""


# ─── JavaScript Generation ───────────────────────────────────────────────────

def js_value(val: Any) -> str:
    """Serialize a Python value to JavaScript literal syntax."""
    if val is None:
        return "null"
    if isinstance(val, bool):
        return "true" if val else "false"
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val == int(val) and abs(val) < 1e15:
            return str(int(val))
        return str(val)
    # String — single-quote with escaping
    s = str(val).replace("\\", "\\\\").replace("'", "\\'").replace("\n", "\\n")
    return f"'{s}'"


def js_object_oneline(obj: Dict[str, Any], key_order: List[str]) -> str:
    """Serialize a dict as a one-line JS object literal with unquoted keys."""
    parts = []
    for key in key_order:
        if key in obj:
            parts.append(f"{key}: {js_value(obj[key])}")
    # Include any extra keys not in the order
    for key in obj:
        if key not in key_order:
            parts.append(f"{key}: {js_value(obj[key])}")
    return "{ " + ", ".join(parts) + " }"


def generate_js_data(data: Dict[str, Any]) -> str:
    """Generate the full defaultRawData JavaScript declaration."""
    lines = ["const defaultRawData = {"]

    # publishers
    if "publishers" in data:
        key_order = ["id", "name", "title", "type", "contact", "renewalDate", "status", "savingsAmount", "savingsType"]
        lines.append("    publishers: [")
        for pub in data["publishers"]:
            lines.append(f"        {js_object_oneline(pub, key_order)},")
        lines.append("    ],")

    # spendData
    if "spendData" in data:
        key_order = ["publisher", "companySpend", "msdSpend", "tiamSpend", "fiscalYear", "notes"]
        lines.append("    spendData: [")
        for s in data["spendData"]:
            lines.append(f"        {js_object_oneline(s, key_order)},")
        lines.append("    ],")

    # riskData
    if "riskData" in data:
        key_order = ["publisher", "sspa", "po", "finance", "legal", "inventory", "details"]
        lines.append("    riskData: [")
        for r in data["riskData"]:
            lines.append(f"        {js_object_oneline(r, key_order)},")
        lines.append("    ],")

    # managedTitles
    if "managedTitles" in data:
        key_order = ["title", "publisher", "category", "licenseCount", "notes"]
        lines.append("    managedTitles: [")
        for t in data["managedTitles"]:
            lines.append(f"        {js_object_oneline(t, key_order)},")
        lines.append("    ],")

    # datasetVersion
    version = data.get("datasetVersion", f"FY26_EXCEL_IMPORT_{date.today().isoformat()}")
    lines.append(f"    datasetVersion: {js_value(version)},")

    # externalKpis
    if "externalKpis" in data:
        key_order = ["name", "value", "unit", "source", "lastUpdated", "notes"]
        lines.append("    externalKpis: [")
        for k in data["externalKpis"]:
            lines.append(f"        {js_object_oneline(k, key_order)},")
        lines.append("    ]")

    lines.append("};")
    return "\n".join(lines)


# ─── data.js Update ──────────────────────────────────────────────────────────

def update_data_js(data_js_path: Path, new_data: Dict[str, Any]) -> str:
    """Replace the defaultRawData object in data.js and return the new content."""
    content = data_js_path.read_text(encoding="utf-8")

    # Find the defaultRawData declaration and its closing
    pattern = re.compile(
        r"(//[^\n]*\n)*\s*const\s+defaultRawData\s*=\s*\{",
        re.MULTILINE
    )
    match = pattern.search(content)
    if not match:
        raise RuntimeError("Could not find 'const defaultRawData = {' in data.js")

    start = match.start()

    # Find the matching closing '};' by counting braces
    brace_count = 0
    end = -1
    in_string = False
    string_char = None
    i = match.end() - 1  # Start at the opening brace

    for idx in range(i, len(content)):
        c = content[idx]

        if in_string:
            if c == "\\" and idx + 1 < len(content):
                continue  # skip escaped chars (handled by next iteration)
            if c == string_char and (idx == 0 or content[idx - 1] != "\\"):
                in_string = False
            continue

        if c in ("'", '"', '`'):
            in_string = True
            string_char = c
            continue

        if c == "{":
            brace_count += 1
        elif c == "}":
            brace_count -= 1
            if brace_count == 0:
                # Look for the semicolon after closing brace
                rest = content[idx + 1:idx + 5]
                semi = rest.find(";")
                end = idx + 1 + semi + 1 if semi >= 0 else idx + 2
                break

    if end == -1:
        raise RuntimeError("Could not find the end of defaultRawData object in data.js")

    # Preserve any leading comments
    leading = content[:start]
    # Find the comment block right before defaultRawData
    comment_lines = []
    pre_lines = leading.rstrip().split("\n")
    for line in reversed(pre_lines):
        stripped = line.strip()
        if stripped.startswith("//"):
            comment_lines.insert(0, line)
        elif stripped == "":
            comment_lines.insert(0, line)
        else:
            break

    comment_block = "\n".join(comment_lines).rstrip()
    pre_content = "\n".join(pre_lines[:len(pre_lines) - len(comment_lines)]) if comment_lines else leading.rstrip()

    new_declaration = generate_js_data(new_data)

    # Rebuild: everything before comments + comments + new data + everything after
    if comment_block.strip():
        new_content = pre_content + "\n" + comment_block + "\n" + new_declaration + content[end:]
    else:
        new_content = content[:start] + new_declaration + content[end:]

    return new_content


# ─── Template Generation ─────────────────────────────────────────────────────

def create_template(output_path: str):
    """Generate a template Excel workbook with the expected sheet/column structure."""
    wb = openpyxl.Workbook()

    # Publishers sheet
    ws = wb.active
    ws.title = "Publishers"
    ws.append(["ID", "Publisher Name", "Product Title", "License Type", "Contact",
               "Renewal Date", "Status", "Savings Amount", "Savings Type"])
    ws.append([1, "Example Publisher", "Example Product", "SaaS", "Contact Name",
               "2026-06-30", "Active", 0, "Cost Avoidance"])

    # Spend sheet
    ws2 = wb.create_sheet("Spend")
    ws2.append(["Publisher", "Company Spend", "MSD Spend", "TI&M Spend", "Fiscal Year", "Notes"])
    ws2.append(["Example Publisher", 1000000, 50000, 10000, "FY26", ""])

    # Risks sheet
    ws3 = wb.create_sheet("Risks")
    ws3.append(["Publisher", "SSPA", "PO", "Finance", "Legal", "Inventory", "Details"])
    ws3.append(["Example Publisher", "", "", "", "", "", ""])

    # Managed Titles sheet
    ws4 = wb.create_sheet("ManagedTitles")
    ws4.append(["Title", "Publisher", "Category", "License Count", "Notes"])
    ws4.append(["Example Product", "Example Publisher", "Other", 0, "active"])

    # External KPIs sheet
    ws5 = wb.create_sheet("ExternalKPIs")
    ws5.append(["KPI Name", "Value", "Unit", "Source", "Last Updated", "Notes"])
    ws5.append(["SNOW Tickets MTD", 0, "tickets", "ServiceNow", "2026-01-01", ""])
    ws5.append(["ICM Tickets MTD", 0, "tickets", "ICM System", "2026-01-01", ""])

    # Auto-size columns
    for ws_item in wb.worksheets:
        for col in ws_item.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws_item.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(output_path)
    print(f"Template saved to: {output_path}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Import SLS MBR data from Excel into data.js",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python import_from_excel.py --file data.xlsx
  python import_from_excel.py --file data.xlsx --dry-run
  python import_from_excel.py --file data.xlsx --merge
  python import_from_excel.py --template sls_template.xlsx
  python import_from_excel.py --file data.xlsx --sheet-map "Sheet1=publishers" "Financials=spend"
        """,
    )
    parser.add_argument("--file", "-f", help="Path to Excel workbook (.xlsx)")
    parser.add_argument("--data-js", default="data.js", help="Path to data.js (default: data.js)")
    parser.add_argument("--dry-run", action="store_true", help="Preview changes without modifying data.js")
    parser.add_argument("--merge", action="store_true",
                        help="Merge with existing data (update matching publishers, add new ones). "
                             "Default: replace all data from Excel.")
    parser.add_argument("--sheet-map", nargs="*", metavar="SHEET=TYPE",
                        help="Explicit sheet-to-type mapping (e.g., 'Sheet1=publishers' 'Financials=spend'). "
                             "Types: publishers, spend, risks, titles, kpis")
    parser.add_argument("--template", metavar="OUTPUT.xlsx",
                        help="Generate a template Excel file and exit")
    args = parser.parse_args()

    # Template mode
    if args.template:
        create_template(args.template)
        return

    if not args.file:
        parser.error("--file is required (or use --template to generate a template)")

    file_path = Path(args.file)
    if not file_path.exists():
        print(f"ERROR: File not found: {file_path}", file=sys.stderr)
        sys.exit(1)

    data_js_path = Path(args.data_js)
    if not data_js_path.exists():
        print(f"ERROR: data.js not found at: {data_js_path}", file=sys.stderr)
        sys.exit(1)

    # Parse explicit sheet mappings
    explicit_map: Dict[str, str] = {}
    if args.sheet_map:
        for mapping in args.sheet_map:
            if "=" not in mapping:
                print(f"ERROR: Invalid sheet-map format (use SHEET=TYPE): {mapping}", file=sys.stderr)
                sys.exit(1)
            sheet_name, sheet_type = mapping.split("=", 1)
            if sheet_type not in SHEET_TYPE_MAPPINGS:
                print(f"ERROR: Unknown type '{sheet_type}'. Valid: {', '.join(SHEET_TYPE_MAPPINGS.keys())}", file=sys.stderr)
                sys.exit(1)
            explicit_map[sheet_name.strip()] = sheet_type.strip()

    # Read workbook — detect format from file header
    print(f"Reading: {file_path}")
    is_xls = _is_old_xls_format(file_path)

    if is_xls:
        if not xlrd:
            print("ERROR: This is an old-format .xls file. Install xlrd: pip install xlrd", file=sys.stderr)
            sys.exit(1)
        print("  (Detected old .xls format — using xlrd)")
        sheet_names, sheet_rows_map = _read_xls(file_path)
    else:
        if not openpyxl:
            print("ERROR: This is an .xlsx file. Install openpyxl: pip install openpyxl", file=sys.stderr)
            sys.exit(1)
        sheet_names, sheet_rows_map = _read_xlsx(file_path)

    imported: Dict[str, Any] = {}
    total_records = 0

    for sheet_name in sheet_names:
        # Determine sheet type
        if sheet_name in explicit_map:
            sheet_type = explicit_map[sheet_name]
        else:
            sheet_type = detect_sheet_type(sheet_name)

        if not sheet_type:
            print(f"  ⚠ Skipping sheet '{sheet_name}' (could not detect type — use --sheet-map)")
            continue

        col_defs = SHEET_TYPE_MAPPINGS[sheet_type]
        rows = sheet_rows_map[sheet_name]
        records, col_mapping, unmapped = read_sheet_data_from_rows(rows, col_defs)

        # Map sheet type to data.js key
        data_key_map = {
            "publishers": "publishers",
            "spend": "spendData",
            "risks": "riskData",
            "titles": "managedTitles",
            "kpis": "externalKpis",
        }
        data_key = data_key_map[sheet_type]

        mapped_cols = {idx: name for idx, name in col_mapping.items()}
        print(f"  ✓ Sheet '{sheet_name}' → {data_key}: {len(records)} records")
        print(f"    Mapped columns: {', '.join(mapped_cols.values())}")
        if unmapped:
            print(f"    ⚠ Unmapped columns (skipped): {', '.join(unmapped)}")

        imported[data_key] = records
        total_records += len(records)

    if not imported:
        print("\nNo data sheets detected. Check sheet names or use --sheet-map.")
        sys.exit(1)

    # Auto-assign IDs if publishers don't have them
    if "publishers" in imported:
        for i, pub in enumerate(imported["publishers"]):
            if not pub.get("id"):
                pub["id"] = i + 1

    # Handle merge mode
    if args.merge:
        # Load existing data
        existing_content = data_js_path.read_text(encoding="utf-8")
        # Parse existing defaultRawData by loading it through the generated JS file approach
        # For simplicity, read the existing data from the generated file if it exists
        existing = _parse_existing_data(data_js_path)
        if existing:
            imported = _merge_data(existing, imported)
            print(f"\n  Merged with existing data")

    # Set dataset version
    imported["datasetVersion"] = f"FY26_EXCEL_IMPORT_{date.today().isoformat()}"

    # If some sections weren't in the Excel, keep existing ones
    if not args.merge:
        _fill_missing_sections(imported, data_js_path)

    print(f"\nTotal: {total_records} records imported across {len(imported) - 1} sections")

    # Generate output
    new_content = update_data_js(data_js_path, imported)

    if args.dry_run:
        print("\n--- DRY RUN (data.js preview) ---")
        # Show just the defaultRawData portion
        match = re.search(r"const defaultRawData = \{", new_content)
        if match:
            preview = new_content[match.start():match.start() + 2000]
            print(preview)
            if len(new_content) - match.start() > 2000:
                print(f"... ({len(new_content) - match.start() - 2000} more characters)")
        print("\nDry run complete. No files were modified.")
        return

    # Write
    data_js_path.write_text(new_content, encoding="utf-8")
    print(f"\n✓ Updated {data_js_path} successfully")
    print(f"  Dataset version: {imported['datasetVersion']}")
    print(f"  Tip: Open index.html in a browser to see the updated dashboard")


def _fill_missing_sections(imported: Dict, data_js_path: Path):
    """If a section wasn't imported from Excel, try to keep the existing one."""
    existing = _parse_existing_data(data_js_path)
    if not existing:
        return
    for key in ["publishers", "spendData", "riskData", "managedTitles", "externalKpis"]:
        if key not in imported and key in existing:
            imported[key] = existing[key]


def _parse_existing_data(data_js_path: Path) -> Optional[Dict]:
    """Best-effort parse of existing defaultRawData from data.js using Node.js."""
    try:
        import subprocess
        # Use Node.js to evaluate the JS and extract the data
        node_script = f"""
        const fs = require('fs');
        const content = fs.readFileSync('{data_js_path.resolve().as_posix()}', 'utf-8');
        // Extract just the defaultRawData
        const match = content.match(/const\\s+defaultRawData\\s*=\\s*(\\{{[\\s\\S]*?\\n\\}});/);
        if (match) {{
            eval('var result = ' + match[1]);
            console.log(JSON.stringify(result));
        }}
        """
        result = subprocess.run(
            ["node", "-e", node_script],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode == 0 and result.stdout.strip():
            return json.loads(result.stdout.strip())
    except Exception:
        pass
    return None


def _merge_data(existing: Dict, imported: Dict) -> Dict:
    """Merge imported data with existing, updating matches and adding new records."""
    merged = dict(existing)

    if "publishers" in imported:
        existing_pubs = {p["name"]: p for p in existing.get("publishers", [])}
        for pub in imported["publishers"]:
            existing_pubs[pub["name"]] = pub
        merged["publishers"] = list(existing_pubs.values())
        # Re-assign IDs
        for i, pub in enumerate(merged["publishers"]):
            pub["id"] = i + 1

    if "spendData" in imported:
        existing_spend = {s["publisher"]: s for s in existing.get("spendData", [])}
        for spend in imported["spendData"]:
            existing_spend[spend["publisher"]] = spend
        merged["spendData"] = list(existing_spend.values())

    if "riskData" in imported:
        existing_risks = {r["publisher"]: r for r in existing.get("riskData", [])}
        for risk in imported["riskData"]:
            existing_risks[risk["publisher"]] = risk
        merged["riskData"] = list(existing_risks.values())

    if "managedTitles" in imported:
        existing_titles = {(t["title"], t.get("publisher", "")): t for t in existing.get("managedTitles", [])}
        for title in imported["managedTitles"]:
            existing_titles[(title["title"], title.get("publisher", ""))] = title
        merged["managedTitles"] = list(existing_titles.values())

    if "externalKpis" in imported:
        existing_kpis = {k["name"]: k for k in existing.get("externalKpis", [])}
        for kpi in imported["externalKpis"]:
            existing_kpis[kpi["name"]] = kpi
        merged["externalKpis"] = list(existing_kpis.values())

    return merged


if __name__ == "__main__":
    main()
